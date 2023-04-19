import os
import re
from datetime import datetime

import imaplib
import email
import webbrowser
import smtplib
import pickle, shelve
import openpyxl

from tkinter.messagebox import *
from bs4 import BeautifulSoup as bs
from openpyxl import load_workbook
from email.header import decode_header
from email import encoders
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase


def change_date_format(date_string):
    date_list = date_string.split(' ')
    day = date_list[1]
    month = date_list[2]
    year = date_list[3]
    time_list = date_list[4].split(':')
    hour = time_list[0]
    min = time_list[1]
    sec = time_list[2]
    new_date_format = f'{day}_{month}_{year}-{hour}_{min}_{sec}'
    return new_date_format

def find_emails(text):
    emails = re.findall(r'[\w.+-]+@[\w-]+\.[\w.-]+', text)
    if len(emails) != 0:
        email = "".join(emails[0])
        return email

def writing_to_files(answer_file_name, letter, mail):
    try:
        os.mkdir('sending_letter/'+ mail + '/')
    except FileExistsError:
        pass
    for i in range(len(letter)):
        letter[i]+='\n'
    txt = open('sending_letter/'+mail+'.txt', 'w', encoding='utf-8')
    txt.writelines(mail+"\n")
    txt.writelines(answer_file_name+"\n")
    txt.writelines(letter)
    txt.close()

def body_creater(page, br_links, quan_of_lnk):
    datafile = open('works_file/answ_text.dat', 'rb')
    start = pickle.load(datafile)
    midle = pickle.load(datafile)
    fn_for_empt = pickle.load(datafile)
    two_links_st = pickle.load(datafile)
    two_links_fn = pickle.load(datafile)
    end = pickle.load(datafile)
    if str(quan_of_lnk) == "missing brocken links":
        body = start[0]+page+midle[0]+br_links+fn_for_empt[0]+end[0]
    elif str(quan_of_lnk) == "1":
        body = start[0]+page+midle[0]+br_links+end[0]
    elif str(quan_of_lnk) == "2":
        body = start[0]+page+two_links_st[0]+br_links.replace("}, {", two_links_fn[0])+end[0]
    return body

def answer(username, password, answer_file_name, letter, mail, from_gmail):
    try:
        print(letter)
        msg = MIMEMultipart("alternative")
        msg["From"] = from_gmail
        msg["To"] = mail
        msg["Subject"] = "".join(letter[2])
        text = answer_file_name
        text_part = MIMEText(text, "plain")
        msg.attach(text_part)
        server = smtplib.SMTP(host="smtp.office365.com", port=587)
        server.starttls()
        server.login(username, password)
        server.sendmail(username, mail, msg.as_string())
        server.quit()
    except:
        print("Invalid To email")

def extr_from_db(gmail):
    list = load_workbook('new_db.xlsx')
    sheet = list.active
    for i in range(1, sheet.max_row):
        if sheet[1+i][0].value != None and sheet[1+i][2].value == gmail and sheet[1+i][6].value != "True" and sheet[1+i][4].value != "Resource not available":
            page = sheet[1+i][1].value
            lxsl_mail = sheet[1+i][2].value
            br_link = sheet[1+i][3].value
            quan_of_lnk = sheet[1+i][4].value
            sheet.cell(1+i, column=7).value = "True"
            break
        else:
            page, lxsl_mail, br_link, quan_of_lnk = False, False, False, False
    list.save('new_db.xlsx')
    return page, lxsl_mail, br_link, quan_of_lnk


def check_keywords(username, password, from_gmail, letter):
    #print(letter)
    key_words = ["Let us know", "send", "to me", "me", "here", "share", "assist", "are you referring to", "assistance", "URL", "let me know", "I can help", "refer", "I help you", "you are looking for", "are you looking for", "take a look", "To this", "elaborate", "show me the links", "identify", "point out", "where", "I am the one", "can you", "what"]
    kew_for_gmail = [" at "]
    mail = None
    for i in key_words:
        if letter[3].count(i) != 0:
            for j in kew_for_gmail:
                if letter[3].count(j) != 0:
                    mail = find_emails(letter[3])
                    if mail == from_gmail:
                        mail = None
                    else:
                        letter[2]=f"from {find_emails(letter[0])} {letter[2]}"
                if mail == None:
                    mail = find_emails(letter[0])
            letter[1]=change_date_format(letter[1])
            page, lxsl_mail, br_links, quan_of_lnk = extr_from_db(find_emails(letter[0]))
            if page != False and page != None:
                body = body_creater(page, br_links, quan_of_lnk)
                #answer file name, letter data, email forwarding
                writing_to_files(body, letter, mail)
                answer(username, password, body, letter, mail, from_gmail)
            break

def read_inbox():
    N=100
    datafile = open('works_file/login_data.dat', 'rb')
    username = pickle.load(datafile)
    password = pickle.load(datafile)
    from_gmail = pickle.load(datafile)
    letter = []
    imap  = imaplib.IMAP4_SSL('outlook.office365.com')
    imap.login(username, password)
    status, messages = imap.select('inbox')
    # total number of emails
    messages = int(messages[0])
    for i in range(messages, messages-N, -1):
        # fetch the email message by ID
        res, msg = imap.fetch(str(i), "(RFC822)")
        for response in msg:
            if isinstance(response, tuple):
                # parse a bytes email into a message object
                msg = email.message_from_bytes(response[1])
                # decode the email subject
                try:
                    subject, encoding = decode_header(msg["Subject"])[0]
                    if isinstance(subject, bytes):
                        subject = subject.decode(encoding)
                except:
                    subject = 'no subject'
                From, encoding = decode_header(msg["From"])[0]
                if isinstance(From, bytes):
                    From = From.decode(encoding)
                Date, encoding = decode_header(msg.get("Date"))[0]
                if isinstance(Date, bytes):
                    Date = Date.decode(encoding)
                letter = [From, Date, "Re:"+subject]
                # if the email message is multipart
                if msg.is_multipart():
                    # iterate over email parts
                    for part in msg.walk():
                        # extract content type of email
                        content_type = part.get_content_type()
                        content_disposition = str(part.get("Content-Disposition"))
                        try:
                            # get the email body
                            body = part.get_payload(decode=True).decode()
                        except:
                            pass
                        if content_type == "text/plain" and "attachment" not in content_disposition:
                            # print text/plain emails and skip attachments
                            letter.append(body)
                            check_keywords(username, password, from_gmail, letter)
    # close the connection and logout
    imap.close()
    imap.logout()

def outlook_mail_list(date_start, date_end):
    datafile = open('works_file/login_data.dat', 'rb')
    username = pickle.load(datafile)
    password = pickle.load(datafile)
    from_gmail = pickle.load(datafile)
    mail = imaplib.IMAP4_SSL('outlook.office365.com')
    mail.login(username, password)
    mail.select('inbox')
    result, data = mail.uid('search', None, '(SINCE {date_start} BEFORE {date_end})'.format(date_start=date_start, date_end=date_end))
    data_list = data[0].split()
    for num in data_list:
        result, msg = mail.uid('fetch', num, '(RFC822)')
        for response in msg:
            if isinstance(response, tuple):
                # parse a bytes email into a message object
                msg = email.message_from_bytes(response[1])
                # decode the email subject
                try:
                    subject, encoding = decode_header(msg["Subject"])[0]
                    if isinstance(subject, bytes):
                        subject = subject.decode(encoding)
                except:
                    subject = 'no subject'
                From, encoding = decode_header(msg["From"])[0]
                if isinstance(From, bytes):
                    From = From.decode(encoding)
                Date, encoding = decode_header(msg.get("Date"))[0]
                if isinstance(Date, bytes):
                    Date = Date.decode(encoding)
                letter = [From, Date, "Re:"+subject]
                # if the email message is multipart
                if msg.is_multipart():
                    # iterate over email parts
                    for part in msg.walk():
                        # extract content type of email
                        content_type = part.get_content_type()
                        content_disposition = str(part.get("Content-Disposition"))
                        try:
                            # get the email body
                            body = part.get_payload(decode=True).decode()
                        except:
                            pass
                        if content_type == "text/plain" and "attachment" not in content_disposition:
                            # print text/plain emails and skip attachments
                            letter.append(body)
                            check_keywords(username, password, from_gmail, letter)
    # close the connection and logout
    mail.close()
    mail.logout()

# start_time = datetime.now()
#body_creater("https://saofrancisco.consuladoportugal.mne.gov.pt/en/about-portugal/useful-links", "https://www.bportugal.pt/en-US/': 'Bank of Portugal'}]", "1")

#read_inbox()
# #outlook_mail_list('20-Mar-2023', '27-Mar-2023')
#
# print("absolute time: ", datetime.now() - start_time)
