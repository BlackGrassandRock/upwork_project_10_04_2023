import time

import re
import openpyxl
import requests

from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor
from concurrent.futures import as_completed

class Controller:
    link_cell_list = []


    def __init__ (self, gmail_column, link_column, start_line, nm_of_cells, nm_of_flows, file):
        self.gmail_column = gmail_column
        self.link_column = link_column
        self.start_line = start_line
        self.nm_of_cells = nm_of_cells
        self.nm_of_flows = nm_of_flows
        self.file = file

    def _check_link(self, br_link):
        if br_link != None and br_link.startswith("http"):
            return True

    def _check_mail(self, mail):
        if mail != None and re.match(r'[^@]+@[^@]+\.[^@]+', mail):
            return True

    #Allocates tasks to threads
    def breakdown_list(self):
        i = 0
        n_queues = (len(self.link_cell_list)+self.nm_of_flows-1)//self.nm_of_flows
        for i in range(i, n_queues):
            list_to_bd  = []
            with ThreadPoolExecutor(self.nm_of_flows) as executor:
                results = executor.map(self.find_broken_links, self.link_cell_list[:self.nm_of_flows])
                for res, mail in enumerate(results):
                    # get the url for the future
                    url = self.link_cell_list[res]
                    if mail == "Resource not available":
                        stat = "Resource not available"
                    elif isinstance(mail, str):
                        stat = "missing brocken links"
                    elif len(mail) == 1 and isinstance(mail, list):
                        stat = "1"
                    elif len(mail) >= 2 and isinstance(mail, list):
                        stat = "2"
                    if url != None and mail != None:
                        list_to_bd.append([url, mail, stat])
                self.writing_to_xlsx(list_to_bd)
            del self.link_cell_list[:self.nm_of_flows]

    def quotes(self, string):
        return string.replace("[{'", "").replace("'}]", "").replace("': '", " ")

    def writing_to_xlsx(self, list_to_bd):
        row = 1
        x = True
        wb = load_workbook('new_db.xlsx')
        ws = wb.active
        while x == True:
            if ws.cell(row=row, column=1).value is None:
                x = False
            else:
                row += 1
        for i in range(len(list_to_bd)):
            for key, value in list_to_bd[i][0].items():
                ws.cell(row+i, column=1).value = key.split('/')[2] #site
                ws.cell(row+i, column=2).value = key #link
                ws.cell(row+i, column=3).value = value #email
                ws.cell(row+i, column=4).value = self.quotes(str(list_to_bd[i][1])) #brocken links
                ws.cell(row+i, column=5).value = list_to_bd[i][2] #status
                ws.cell(row+i, column=6).value = 0
                ws.cell(row+i, column=7).value = 0
        wb.save('new_db.xlsx')

    def extract_data_from_xlsx(self):
        i = 0
        list = openpyxl.open(self.file+".xlsx", read_only=True)
        sheet = list.active
        for i in range(i, self.nm_of_cells):
            br_link = sheet[self.start_line+i][self.link_column].value
            mail = sheet[self.start_line+i][self.gmail_column].value
            if self._check_link(br_link) and self._check_mail(mail):
                self.link_cell_list.append({br_link:mail})

    def checker(self, response):
            if response.status_code in [404, 410, 400, 406]:
                return True

    def find_broken_links(self, dict_url):
        for url, _ in dict_url.items():
            pass
        substitution = "Resource not available"
        broken_links = []
        timeout = True
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; rv:91.0) Gecko/20100101 Firefox/91.0'}
        print(url)
        try:
            page = requests.get(url, headers=headers, timeout=(1, 2))
        except:
            timeout = False
        if timeout and self.checker(page) != True:
            links = BeautifulSoup(page.content, 'html.parser').select("a[href]")
            links = set(links)
            for link in links:
                link_url = link.get('href')
                #Checking if the link works
                if link_url.startswith("https") and "home" not in link_url and "default" not in link_url:
                    substitution = link_url
                    text = re.sub("^\s+|\n|\r|\s+$", '', link.text)
                    try:
                        response = requests.head(link_url, timeout=1, headers=headers)
                    except:
                        timeout = False
                    if timeout and self.checker(response):
                        broken_links.append({link_url:text})
            try:
                broken_links = list(set(broken_links))
            except:
                pass
            if broken_links == []:
                return str(substitution)
            else:
                return broken_links
        else:
            return "Resource not available"


def br_ln(gmail_column, link_column, start_line, nm_of_cells, nm_of_flows, file):
    print(gmail_column, link_column, start_line, nm_of_cells, nm_of_flows, file)
    settings = Controller(gmail_column, link_column, start_line, nm_of_cells, nm_of_flows, file)
    settings.extract_data_from_xlsx()
    settings.breakdown_list()
