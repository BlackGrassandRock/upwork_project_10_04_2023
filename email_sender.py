import smtplib
import os
import time
import pickle, shelve
import openpyxl

from tkinter.messagebox import *
from openpyxl import load_workbook
from email import encoders
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from bs4 import BeautifulSoup as bs


def quotes(string):
    return string.replace('[{', '').replace('}]', '').replace('}, {', '')

def find_file(intermid_file):
    for i in os.listdir('upd_bd/'+intermid_file+'/'):
        return 'upd_bd/'+intermid_file+'/'+i

def answer(username, password, from_gmail, to_mail, subject, mess):
    try:
        to_mail="alexghfs915@gmail.com"
        msg = MIMEMultipart("alternative")
        msg["From"] = from_gmail
        msg["To"] = to_mail
        msg["Subject"] = subject
        text = mess
        text_part = MIMEText(text, "plain")
        msg.attach(text_part)
        server = smtplib.SMTP(host="smtp.office365.com", port=587)
        server.starttls()
        server.login(username, password)
        server.sendmail(username, to_mail, msg.as_string())
        server.quit()
    except:
        print("Invalid To email")

def sender_controller(delay_time):
    start_string = 1
    counter = 0
    sbj = "Unable to Open Some of the Links on "
    text = ["Hi!\nI was on your site ", " and came across a few links that were not working.\n\nDo you mind telling me who I should send them over to?\n\nThanks! :)"]
    datafile = open('works_file/login_data.dat', 'rb')
    username = pickle.load(datafile)
    password = pickle.load(datafile)
    from_gmail = pickle.load(datafile)
    print(username, password, from_gmail)

    list = load_workbook('new_db.xlsx')
    sheet = list.active
    n = sheet.max_row

    try:
        for i in range(n):
            if sheet[start_string+i][5].value != "True" and sheet[1+i][4].value != "Resource not available" and sheet[start_string+i][0].value != None:
                site = sheet[start_string+i][0].value
                mail = sheet[start_string+i][2].value
                sheet.cell(start_string+i, column=6).value = "True"
                subject = sbj+site
                mess = text[0]+site+text[1]
                answer(username, password, from_gmail, mail, subject, mess)
                print(username, password, from_gmail, mail, subject, mess)
                time.sleep(delay_time)
                counter+=1
        list.save('new_db.xlsx')
        showinfo("Отправка нульовок", f"было отправлено {counter} писем\n")
    except:
        showinfo("Отправка нульовок", f"Отправка завершина.")
